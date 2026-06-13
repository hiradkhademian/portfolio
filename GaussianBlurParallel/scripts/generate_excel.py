#!/usr/bin/env python3
"""
Generate Excel file from benchmark CSV results
"""

import csv
import sys
from datetime import datetime

def generate_excel_from_csv():
    """Generate Excel file from combined benchmark CSV"""
    
    csv_file = "GaussianBlur_Combined_Results.csv"
    xlsx_file = "GaussianBlur_Benchmark_Analysis.xlsx"
    
    try:
        # Try to use openpyxl if available
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Benchmark Results"
            
            # Read CSV
            with open(csv_file, 'r') as f:
                reader = csv.reader(f)
                rows = list(reader)
            
            # Filter data rows
            data_rows = []
            summary_data = {}
            in_summary = False
            
            for row in rows:
                if not row or row[0] == "":
                    continue
                if row[0] == "Metric":
                    in_summary = True
                if in_summary:
                    if row[0] != "Metric" and row[0] != "":
                        summary_data[row[0]] = row[1] if len(row) > 1 else ""
                elif row[0].startswith("Image File") or row[0].startswith("="):
                    continue
                else:
                    data_rows.append(row)
            
            # Write header
            header = ["Image File", "Width", "Height", "Total Pixels", "Seq Min (ms)", "Seq Max (ms)", 
                     "Seq Avg (ms)", "Seq Median (ms)", "FJ Min (ms)", "FJ Max (ms)", "FJ Avg (ms)", 
                     "FJ Median (ms)", "Speedup", "Efficiency (%)", "Cores"]
            ws.append(header)
            
            # Style header
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data
            for row in data_rows:
                if len(row) >= 15 and (row[0].endswith(".jpeg") or row[0].endswith(".png")):
                    ws.append(row)
            
            # Add summary section
            summary_row = len(data_rows) + 3
            ws.cell(row=summary_row, column=1, value="SUMMARY STATISTICS")
            ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
            
            summary_row += 1
            for metric, value in summary_data.items():
                if metric and value:
                    ws.cell(row=summary_row, column=1, value=metric)
                    ws.cell(row=summary_row, column=2, value=value)
                    summary_row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save
            wb.save(xlsx_file)
            print(f"✅ Excel file created: {xlsx_file}")
            return True
            
        except ImportError:
            print("⚠️  openpyxl not available. Generating CSV-based report instead...")
            
            # Fallback: Generate formatted CSV
            with open(csv_file, 'r') as infile:
                content = infile.read()
            
            # Read and reformat
            with open("GaussianBlur_Benchmark_Results_Formatted.csv", 'w') as outfile:
                outfile.write(content)
            
            print(f"✅ Formatted CSV created: GaussianBlur_Benchmark_Results_Formatted.csv")
            print("💡 You can open this CSV directly in Excel")
            return True
            
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

if __name__ == "__main__":
    print("=" * 50)
    print("  GENERATING EXCEL REPORT")
    print("=" * 50)
    print()
    
    success = generate_excel_from_csv()
    
    if success:
        print()
        print("✅ Report generation complete!")
        print()
        print("Files available for analysis:")
        print("  • GaussianBlur_Benchmark_Analysis.xlsx (Excel file)")
        print("  • GaussianBlur_Combined_Results.csv (CSV format)")
        print("  • GaussianBlur_Comprehensive_Report.html (HTML report)")
    else:
        sys.exit(1)
