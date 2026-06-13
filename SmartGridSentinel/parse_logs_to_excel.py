#!/usr/bin/env python3
"""
Parse docker compose logs and export to Excel with formatted structure
"""

import re
import pandas as pd
from datetime import datetime
from pathlib import Path

# Configuration
LOG_FILE = "system_logs.txt"
EXCEL_FILE = "grid_sentinel_logs.xlsx"
ROOT_DIR = Path(__file__).parent

class LogParser:
    def __init__(self, log_file):
        self.log_file = ROOT_DIR / log_file
        self.events = []
        
    def parse(self):
        """Parse log file and extract events"""
        if not self.log_file.exists():
            print(f"❌ Log file not found: {self.log_file}")
            return []
        
        with open(self.log_file, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        
        for line in lines:
            event = self._extract_event(line)
            if event:
                self.events.append(event)
        
        print(f"✓ Parsed {len(self.events)} events from logs (4 core services only)")
        return self.events
    
    def _extract_event(self, line):
        """Extract structured event from a log line"""
        
        # Skip empty lines
        if not line.strip():
            return None
        
        # Parse timestamp (look for HH:MM:SS pattern)
        timestamp_match = re.search(r'(\d{2}:\d{2}:\d{2})', line)
        if not timestamp_match:
            return None
        
        timestamp = timestamp_match.group(1)
        
        # Identify service source
        service = self._identify_service(line)
        
        # Identify target (meter or zone)
        target = self._identify_target(line)
        
        # Identify triggering event
        trigger = self._identify_trigger(line)
        
        # Identify action taken
        action = self._identify_action(line)
        
        # Identify result
        result = self._identify_result(line)
        
        # Identify DLQ status
        dlq_status = self._identify_dlq_status(line)
        
        # Only return if we have meaningful content AND valid service
        if service and (trigger or action or dlq_status):
            return {
                "Zaman Damgası": timestamp,
                "Kaynak Servis": service,
                "Hedef (Cihaz/Bölge)": target,
                "Tetikleyici Olay": trigger,
                "Alınan Aksiyon": action,
                "Sonuç / Sistem Durumu": result,
                "DLQ Durumu": dlq_status
            }
        
        return None
    
    def _identify_service(self, line):
        """Identify which service generated the log"""
        if "ingestion" in line.lower():
            return "Ingestion Service"
        elif "real-time-analysis" in line.lower() or "real_time_analysis" in line.lower():
            return "Real-Time Analysis"
        elif "trend-regional" in line.lower() or "trend_regional" in line.lower():
            return "Trend & Regional Analysis"
        elif "action-gateway" in line.lower() or "action_gateway" in line.lower():
            return "Action Gateway"
        elif "mock-engine" in line.lower() or "mock_engine" in line.lower():
            return None  # Filtered out
        elif "dlq-monitor" in line.lower() or "dlq_monitor" in line.lower():
            return None  # Filtered out
        else:
            return None  # Skip unknown services
    
    def _identify_target(self, line):
        """Identify target device or zone"""
        # Meter patterns
        meter_match = re.search(r'(METER-\d{2}[A-Z])', line)
        if meter_match:
            return meter_match.group(1)
        
        # Zone patterns
        zone_match = re.search(r'(ZONE-[A-Z]+)', line)
        if zone_match:
            return zone_match.group(1)
        
        # Regional gateway
        if "REGIONAL-GATEWAY" in line:
            return "REGIONAL-GATEWAY"
        
        return "Şebeke Geneli"  # Network-wide
    
    def _identify_trigger(self, line):
        """Identify triggering event"""
        triggers = {
            "VoltageSpikeDetected": "Voltage Spike Anomaly",
            "CurrentSpikeDetected": "Current Spike Anomaly",
            "PowerSpikeDetected": "Power Spike Anomaly",
            "BlackoutDetected": "Blackout Detected",
            "SuspiciousConsumption": "Suspicious Consumption Pattern",
            "Rapid load growth": "Rapid Load Growth",
            "CHAOS TEST": "Chaos Test (Corrupted Packet)",
            "Auto-recovery": "Auto-recovery Timer",
            "Throttle released": "Throttle Release",
            "ANOMALY DETECTED": "Anomaly Detected",
        }
        
        for key, value in triggers.items():
            if key.lower() in line.lower():
                return value
        
        # Voltage/Current specific patterns
        if re.search(r'(\d{3}\.\d+V)', line):
            voltage_match = re.search(r'(\d{3}\.\d+V)', line)
            if voltage_match:
                return f"Voltage Spike ({voltage_match.group(1)})"
        
        if re.search(r'(\d{2}\.\d+A)', line):
            current_match = re.search(r'(\d{2}\.\d+A)', line)
            if current_match:
                return f"Current Spike ({current_match.group(1)})"
        
        return None
    
    def _identify_action(self, line):
        """Identify action taken"""
        actions = {
            "CUT_POWER": "CUT_POWER Command",
            "RESTART_METER": "RESTART_METER Command",
            "THROTTLE_CONSUMPTION": "THROTTLE_CONSUMPTION Command",
            "DLQ Isolation": "DLQ Isolation (Rejected)",
            "Emergency Alert Dispatch": "Emergency Alert Dispatch",
            "ACK Received": "Command Acknowledged (ACK)",
            "Audit log recorded": "Audit Log Recorded",
            "Power restored": "Power Restored",
            "Power physically cut": "Power Physically Cut",
        }
        
        for key, value in actions.items():
            if key.lower() in line.lower():
                return value
        
        return None
    
    def _identify_result(self, line):
        """Identify result/system status"""
        results = {
            "ACK Received": "Komut başarıyla iletildi (ACK) ve veritabanına denetim logu kaydedildi.",
            "Power restored": "Şebeke gücü otonom olarak geri getirildi.",
            "Power physically cut": "Donanım gücü fiziksel olarak kesildi. Otonom onarım başlatıldı.",
            "successfully isolated": "Bozuk formatlı veri DLQ altyapısına izole edildi.",
            "Throttle released": "Bölgesel yük hafifledi, cihaz normal tüketime döndürüldü.",
        }
        
        for key, value in results.items():
            if key.lower() in line.lower():
                return value
        
        # Generic success/failure patterns
        if "✅" in line or "success" in line.lower():
            return "İşlem başarıyla tamamlandı"
        elif "❌" in line or "error" in line.lower():
            return "Hata oluştu"
        
        return "İşlem gerçekleştirildi"
    
    def _identify_dlq_status(self, line):
        """Identify DLQ (Dead Letter Queue) status and type"""
        dlq_types = {
            "telemetry-dlq": "DLQya alindi (Telemetry)",
            "telemetry_dlq": "DLQya alindi (Telemetry)",
            "emergency-alerts-dlq": "DLQya alindi (Emergency Alert)",
            "emergency_alerts_dlq": "DLQya alindi (Emergency Alert)",
            "trend-region-dlq": "DLQya alindi (Trend Region)",
            "trend_region_dlq": "DLQya alindi (Trend Region)",
            "action-gateway-dlq": "DLQya alindi (Action Gateway)",
            "action_gateway_dlq": "DLQya alindi (Action Gateway)",
            "dead_letter": "DLQya alindi (Unknown)",
            "dead-letter": "DLQya alindi (Unknown)",
            "dlq": "DLQya alindi (Unknown)",
        }
        
        for key, value in dlq_types.items():
            if key.lower() in line.lower():
                return value
        
        # Check for DLQ-specific keywords
        if "isolated" in line.lower() and ("dlq" in line.lower() or "dead" in line.lower()):
            return "DLQya alindi (Isolated)"
        
        return None

def create_excel(events, output_file):
    """Create professionally formatted Excel file from events"""
    if not events:
        print("⚠️  No events to export")
        return False
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    # Create DataFrame
    df = pd.DataFrame(events)
    
    # Reorder columns - INCLUDING new DLQ Durumu column
    columns = ["Zaman Damgası", "Kaynak Servis", "Hedef (Cihaz/Bölge)", 
               "Tetikleyici Olay", "Alınan Aksiyon", "Sonuç / Sistem Durumu", "DLQ Durumu"]
    
    # Ensure all columns exist, fill missing with None
    for col in columns:
        if col not in df.columns:
            df[col] = None
    
    df = df[columns]
    
    # Remove duplicates
    df = df.drop_duplicates()
    
    # Sort by timestamp
    df = df.sort_values("Zaman Damgası")
    
    # Write to Excel
    output_path = ROOT_DIR / output_file
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Grid Sentinel Logs', index=False, startrow=2)
            
            worksheet = writer.sheets['Grid Sentinel Logs']
            
            # Professional styling configuration
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            
            alt_row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            dlq_highlight = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow for DLQ
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            center_alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Add title
            title_cell = worksheet['A1']
            title_cell.value = f"SmartGrid Sentinel - Event Log Report ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})"
            title_cell.font = Font(name='Calibri', size=14, bold=True, color="1F4E78")
            worksheet.merge_cells('A1:G1')
            title_cell.alignment = center_alignment
            
            # Format header row (row 3 since we have title at row 1)
            for col_num, col_name in enumerate(columns, 1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.value = col_name
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_alignment
            
            # Set column widths
            col_widths = {
                'A': 14,  # Timestamp
                'B': 22,  # Service
                'C': 20,  # Target
                'D': 28,  # Trigger
                'E': 25,  # Action
                'F': 40,  # Result
                'G': 28   # DLQ Status (NEW)
            }
            
            for col_letter, width in col_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # Format data rows
            for row_num, row_data in enumerate(df.values, start=4):
                for col_num, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    cell.alignment = left_alignment if col_num in [2, 3, 4, 5, 6, 7] else center_alignment
                    
                    # Alternate row colors for readability
                    if row_num % 2 == 0:
                        cell.fill = alt_row_fill
                    
                    # Highlight DLQ events in yellow
                    if col_num == 7:  # DLQ Durumu column
                        if value and isinstance(value, str) and "dlq" in value.lower():
                            cell.fill = dlq_highlight
                            cell.font = Font(name='Calibri', size=10, bold=True, color="C65911")  # Orange text
                    
                    # Conditional formatting for status column
                    if col_num == 6:  # Result column
                        if value and isinstance(value, str):
                            if "başarıyla" in value.lower() or "success" in value.lower():
                                cell.font = Font(name='Calibri', size=10, color="00B050")  # Green
                            elif "error" in value.lower() or "hata" in value.lower():
                                cell.font = Font(name='Calibri', size=10, color="C00000")  # Red
                            else:
                                cell.font = Font(name='Calibri', size=10, color="1F4E78")
            
            # Freeze header rows
            worksheet.freeze_panes = 'A4'
            
            # Set row height for header
            worksheet.row_dimensions[1].height = 25
            worksheet.row_dimensions[3].height = 22
            
            # Add data validation / filtering
            worksheet.auto_filter.ref = f'A3:G{len(df) + 3}'
            
            # Set print options
            worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
            worksheet.page_margins.left = 0.5
            worksheet.page_margins.right = 0.5
            worksheet.print_options.horizontalCentered = False
        
        print(f"✓ Excel file created: {output_path}")
        print(f"📊 Professional formatting applied: Headers, colors, filtering, borders, and DLQ tracking")
        return True
    
    except Exception as e:
        print(f"❌ Error creating Excel: {e}")
        return False

def main():
    """Main execution"""
    print("=" * 60)
    print("SmartGrid Sentinel - Log Parser & Excel Generator")
    print("=" * 60)
    
    # Parse logs
    parser = LogParser(LOG_FILE)
    events = parser.parse()
    
    if not events:
        print("⚠️  No valid events found in logs")
        return
    
    # Create Excel
    success = create_excel(events, EXCEL_FILE)
    
    if success:
        print(f"📊 Total events logged: {len(events)}")
        print(f"✓ Export complete: {ROOT_DIR / EXCEL_FILE}")
    else:
        print("❌ Export failed")

if __name__ == "__main__":
    main()
